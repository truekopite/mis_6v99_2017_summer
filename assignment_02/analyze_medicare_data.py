
# coding: utf-8

# In[1]:

import requests
import os
import zipfile
import openpyxl
import sqlite3
import glob
import getpass
import pandas as pd
from openpyxl import load_workbook


# In[2]:

#The link
link = 'https://data.medicare.gov/views/bg9k-emty/files/0a9879e0-3312-4719-a1db-39fd114890f1?content_type=application%2Fzip%3B%20charset%3Dbinary&filename=Hospital_Revised_Flatfiles.zip'
data = requests.get(link)


# In[3]:

new_dir = "staging_test"
#Making a directory named 'staging_test'
os.mkdir(new_dir)


# In[4]:

#Putting the ZIP file in the directory
medical_file =os.path.join(new_dir,"medicare_hospital_compare.zip")
#Opening the ZIP file
medical = open(medical_file,"wb")
medical.write(data.content)
medical.close()


# In[5]:

#Extracting from ZIP file
med = zipfile.ZipFile(medical_file,'r')
med.extractall(new_dir)
med.close()


# In[6]:

#Create the database
conn = sqlite3.connect('medicare_hospital_compare.db')
#Setting the cursor
c1 = conn.cursor()
#Selectin only .csv files
main_dir = os.path.join(new_dir, "*.csv")
for file_name in glob.glob(main_dir):
    if os.path.basename(file_name) != 'FY2015_Percent_Change_in_Medicare_Payments.csv':
        #Changing directory
        os.chdir(new_dir)
        name = os.path.basename(file_name)
        fn = os.path.join(new_dir,name)
        #Opening the csv file in cp1252 encoding
        in_fp = open(name,"rt",encoding='cp1252')
        input_data = in_fp.read()
        in_fp.close()
        new_name = name + "_fix"
        #Writing the cav files as csv_fix with utf-8 encoding
        out_fp = open(new_name,"wt",encoding='utf-8')
        for c in input_data:
            if c != '\0':
                out_fp.write(c)       
        out_fp.close()
        #Reading csv to dataframe
        df = pd.read_csv(new_name, dtype=str)
        #Changing name formatting
        new_name = new_name.lower()
        new_name = new_name.replace(".csv_fix", "")
        new_name = new_name.replace(" ", "_")
        new_name = new_name.replace("-", "_")
        new_name = new_name.replace("%", "pct")
        new_name = new_name.replace("/", "_")
        if (new_name[0].isalpha() != 1):
            new_name = "t_" +new_name
        #Changing column name formatting
        df.rename(columns = lambda x: x.lower(), inplace = True)
        df.rename(columns = lambda x: x.replace(' ', '_'), inplace = True)
        df.rename(columns = lambda x: x.replace('-', '_'), inplace = True)
        df.rename(columns = lambda x: x.replace('%', 'pct'), inplace = True)
        df.rename(columns = lambda x: x.replace('/', '_'), inplace = True)
        df.rename(columns = lambda x: "c_" + x if x[0].isalpha() != 1 else x, inplace = True)
        #Writing to sql database
        df.to_sql(new_name, conn, if_exists='replace', index = False)
        #Returning to main directory
        os.chdir("../")
    else:
        #If that file name, then ignore
        continue
conn.commit()


# In[7]:

#The URL
hosp_url = "http://kevincrook.com/utd/hospital_ranking_focus_states.xlsx"
data = requests.get(hosp_url)
#Name of the excel file
ranking_xlsx_fname = "hospital_ranking_focus_states.xlsx"
hosp = open(ranking_xlsx_fname,"wb")
hosp.write(data.content)
hosp.close()
#Loading the workbook
wb = openpyxl.load_workbook(ranking_xlsx_fname)
for sheet_name in wb.get_sheet_names():
    #Name changing formatting
    new_sheet_name = sheet_name
    new_sheet_name = new_sheet_name.lower()
    new_sheet_name = new_sheet_name.replace(".csv_fix", "")
    new_sheet_name = new_sheet_name.replace(" ", "_")
    new_sheet_name = new_sheet_name.replace("-", "_")
    new_sheet_name = new_sheet_name.replace("%", "pct")
    new_sheet_name = new_sheet_name.replace("/", "_")
    if (new_sheet_name[0].isalpha() != 1):
        new_sheet_name = "t_" +new_sheet_name
    #Reading excel file into dataframe
    data_sheet = pd.read_excel(ranking_xlsx_fname,sheet_name, dtype="str")
    #Changing column names formatting of the dataframe
    data_sheet.rename(columns = lambda x: x.lower(), inplace = True)
    data_sheet.rename(columns = lambda x: x.replace(' ', '_'), inplace = True)
    data_sheet.rename(columns = lambda x: x.replace('-', '_'), inplace = True)
    data_sheet.rename(columns = lambda x: x.replace('%', 'pct'), inplace = True)
    data_sheet.rename(columns = lambda x: x.replace('/', '_'), inplace = True)
    data_sheet.rename(columns = lambda x: "c_" + x if x[0].isalpha() != 1 else x, inplace = True)
    #Converting daraframes to sql table in database
    data_sheet.to_sql(new_sheet_name, conn, if_exists='replace', index=False)


# In[8]:

#Column name
col_nam = ["Provider ID", "Hospital Name", "City", "State" ,"County" ]
#The query
nation_query = """select rank.provider_id, hospital_name, city, state, county_name
                    from  hospital_national_ranking as rank 
                    left join hospital_general_information as info
                    on rank.provider_id = info.provider_id
                    order by CAST(ranking AS INT)
                    LIMIT 100"""
#Read the sql query
dat_nation = pd.read_sql(sql = nation_query, con = conn)
#Renaming the columns
dat_nation.columns = col_nam
#Excel file name
wb1_name = "hospital_ranking.xlsx"
#Writing the excel file
writer = pd.ExcelWriter(wb1_name, engine='xlsxwriter')
dat_nation.to_excel(writer, sheet_name='Nationwide', index = False)
writer.save()
focus_sheet = pd.read_excel(ranking_xlsx_fname,"Focus States", dtype="str")
#Sort
focus_sheet.sort_values(by = 'State Name', ascending=True)
list_state = focus_sheet.set_index('State Abbreviation').T.to_dict('records')
#From list to dictionary
list_state = list_state[0]
for abv,sname in list_state.items(): 
    state_que = """select rank.provider_id, hospital_name, city, state, county_name
                    from  hospital_national_ranking as rank 
                    left join hospital_general_information as info
                    on rank.provider_id = info.provider_id
                    where state = '"""+abv+ """'
                    order by CAST(ranking AS INT)
                    LIMIT 100"""
    #Real sql query
    dat_nation = pd.read_sql(sql = state_que, con = conn)
    #Renaming columns
    dat_nation.columns = col_nam
    book = load_workbook(wb1_name)
    writer = pd.ExcelWriter(wb1_name, engine='openpyxl') 
    writer.book = book
    #Create excel sheets
    dat_nation.to_excel(writer, sheet_name=sname, index = False)
    writer.save()


# In[9]:

#Column names
col_meas_name = ["Minimum", "Maximum" ,"Average", "Standard Deviation" ]
#The query
meas_query = """select state, measure_id as "Measure ID", measure_name as "Measure Name", CAST(score AS INT) as score ,
                    case when CAST(score AS INT) then 0 else 1 end as score_eval
                    from  timely_and_effective_care___hospital
                    where score_eval = 0
                    order by measure_id"""
#Read sql query
measure_data = pd.read_sql(sql = meas_query, con = conn)
#Creating analysis data
min_measure = measure_data.groupby(["Measure ID","Measure Name"])["score"].min()
max_measure = measure_data.groupby(["Measure ID","Measure Name"])["score"].max()
avg_measure = measure_data.groupby(["Measure ID","Measure Name"])["score"].mean()
std_measure = measure_data.groupby(["Measure ID","Measure Name"])["score"].std()
#Concatinate the columns
measure_data_final = pd.concat([min_measure,max_measure,avg_measure,std_measure], axis = 1)
#Rename the column names
measure_data_final.columns = col_meas_name
#Workbook name
wb2_name = "measures_statistics.xlsx"
writer = pd.ExcelWriter(wb2_name, engine='xlsxwriter')
measure_data_final.to_excel(writer, sheet_name='Nationwide', index = True)
writer.save()
for abv,sname in list_state.items(): 
    #State query
    state_que = """select state, measure_id as "Measure ID", measure_name as "Measure Name", CAST(score AS INT) as score ,
                    case when CAST(score AS INT) then 0 else 1 end as score_eval
                    from  timely_and_effective_care___hospital
                    where score_eval = 0 and state = '"""+abv+ """'
                    order by measure_id"""
    #Reading the sql query
    measure_data = pd.read_sql(sql = state_que, con = conn)
    #Grouping the analysis data
    min_measure = measure_data.groupby(["Measure ID","Measure Name"])["score"].min()
    max_measure = measure_data.groupby(["Measure ID","Measure Name"])["score"].max()
    avg_measure = measure_data.groupby(["Measure ID","Measure Name"])["score"].mean()
    std_measure = measure_data.groupby(["Measure ID","Measure Name"])["score"].std()
    #Concatinating the data
    measure_data_final = pd.concat([min_measure,max_measure,avg_measure,std_measure], axis = 1)
    #Changing column names
    measure_data_final.columns = col_meas_name
    book = load_workbook(wb2_name)
    writer = pd.ExcelWriter(wb2_name, engine='openpyxl') 
    writer.book = book
    #Writing to excel sheet
    measure_data_final.to_excel(writer, sheet_name=sname, index = True)
    writer.save()

